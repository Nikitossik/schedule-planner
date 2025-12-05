from fastapi import HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_

from app.models.schedule import Schedule
from app.repositories.schedule import ScheduleRepository
from app.schemas.schedule import ScheduleIn, ScheduleExportParams

from app.services.base import BaseService

import io
from datetime import datetime, timedelta
from typing import List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# PDF imports
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm

from ..models import Lesson, Group, StudyForm
from app.services.university_holiday import UniversityHolidayService


class ScheduleService(BaseService[Schedule, ScheduleIn]):
    """
    Service layer for schedule management.

    Responsibilities:
    - Create schedules with uniqueness validation per (semester, direction).
    - Export schedules in Excel or PDF formats.
    - Provide helper routines for building tabular representations (time slots, table data, styling).
    """

    def __init__(self, db: Session):
        """
        Initialize the Schedule service.

        Args:
            db (Session): Active SQLAlchemy session.
        """
        super().__init__(db, Schedule, ScheduleRepository(db))
        self.holiday_service = UniversityHolidayService(db)
        self.holiday_service = UniversityHolidayService(db)

    def create(self, schedule: ScheduleIn) -> Schedule:
        """
        Create a new schedule ensuring uniqueness within the semester and direction.

        Args:
            schedule (ScheduleIn): Payload for the schedule to create.

        Returns:
            Schedule: Newly created schedule.

        Raises:
            HTTPException: 400 if a schedule for the same semester and direction already exists.
        """
        found_schedule = (
            self.db.query(Schedule)
            .filter(
                Schedule.semester_id == schedule.semester_id,
                Schedule.direction_id == schedule.direction_id,
            )
            .first()
        )

        if found_schedule:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Schedule for this semester already exists.",
            )

        return super().create(schedule)

    def export(
        self,
        schedule_id: int,
        export_params: ScheduleExportParams,
    ) -> StreamingResponse:
        """
        Export a schedule to the requested format (Excel or PDF) and return a streaming response.

        The method delegates to the specific exporter based on export_params.format,
        builds an appropriate content type, and returns a StreamingResponse with
        download headers.

        Args:
            schedule_id (int): ID of the schedule to export.
            export_params (ScheduleExportParams): Export options (format, group_ids, filename).

        Returns:
            StreamingResponse: Streamed file content with proper headers for download.
        """
        """
        Universal schedule export method with error handling
        Returns a ready StreamingResponse for return from the route
        """
        # Вызываем соответствующий метод экспорта
        if export_params.format == "excel":
            file_buffer, generated_filename = self.export_schedule_excel(
                schedule_id, export_params.group_ids, export_params.filename
            )
            media_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:  # pdf
            file_buffer, generated_filename = self.export_schedule_pdf(
                schedule_id, export_params.group_ids, export_params.filename
            )
            media_type = "application/pdf"

        # Формируем заголовки
        headers = {
            "Content-Disposition": f'attachment; filename="{generated_filename}"',
            "Content-Type": media_type,
        }

        # Возвращаем StreamingResponse
        return StreamingResponse(
            file_buffer,
            media_type=media_type,
            headers=headers,
        )

    def export_schedule_excel(
        self,
        schedule_id: int,
        group_ids: Optional[List[int]] = None,
        custom_filename: Optional[str] = None,
    ) -> Tuple[io.BytesIO, str]:
        """
        Export a schedule to Excel in the tabular layout used by the UI.

        Steps:
        - Load schedule and lessons (optionally filtered by groups).
        - Generate time slots (08:00–18:00, 45-minute slots with 10-minute breaks).
        - Prepare styles and fill the sheet.
        - Return an in-memory buffer and a safe filename.

        Args:
            schedule_id (int): ID of the schedule to export.
            group_ids (Optional[List[int]]): Subset of group IDs to include; None for all.
            custom_filename (Optional[str]): Base filename without extension; defaults to a generated name.

        Returns:
            Tuple[io.BytesIO, str]: (buffer, filename) where buffer is positioned at start.
        """
        """Экспорт расписания в Excel в формате как на картинке"""

        # Получаем расписание
        schedule = self.get_by_id(schedule_id)

        # Получаем уроки для расписания (с фильтрацией по группам если указаны)
        lessons = self._get_lessons_for_schedule(schedule_id, group_ids)

        # Создаем временные слоты (с 8:00 до 18:00 с интервалом 45 минут)
        time_slots = self._generate_time_slots()

        # Создаем Excel файл
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Schedule"

        # Устанавливаем базовый шрифт Arial 9 для всего листа
        workbook.properties.application = "Schedule Planner"

        # УСТАНАВЛИВАЕМ ARIAL 9 КАК БАЗОВЫЙ ШРИФТ ДЛЯ ВСЕГО WORKBOOK
        from openpyxl.styles import NamedStyle

        # Создаем базовый стиль Arial 9
        base_style = NamedStyle(name="base_style")
        base_style.font = Font(name="Arial", size=9, color="000000")

        # Добавляем стиль в workbook
        if "base_style" not in workbook.named_styles:
            workbook.add_named_style(base_style)

        # Устанавливаем стандартную ширину колонок точно 8.5 символов для Arial 9
        worksheet.sheet_format.defaultColWidth = 10
        worksheet.sheet_format.baseColWidth = 10

        # Устанавливаем ширину колонки A (дни недели) - 12 символов
        worksheet.column_dimensions["A"].width = 12

        # Определяем период дат из уроков или семестра
        if lessons:
            start = min(lesson.date for lesson in lessons)
            end = max(lesson.date for lesson in lessons)
        else:
            # Если уроков нет, используем даты семестра
            start = schedule.semester.start_date
            end = schedule.semester.end_date

        # Получаем выходные дни для этого периода
        holiday_service = UniversityHolidayService(self.db)
        holiday_dates_list = holiday_service.get_expanded_holiday_dates(start, end)
        holiday_dates = {holiday["date"] for holiday in holiday_dates_list}

        # Настройка стилей
        self._setup_styles(
            worksheet, schedule, start, end, time_slots, [], lessons, holiday_dates
        )

        # Генерируем имя файла
        if custom_filename:
            filename = f"{custom_filename}.xlsx"
        else:
            filename = f"schedule_{schedule.direction.name}_{schedule.semester.number}_{start.strftime('%Y-%m-%d')}_to_{end.strftime('%Y-%m-%d')}.xlsx"
        filename = filename.replace(" ", "_").replace("/", "_")

        # Сохранение в BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return output, filename

    def export_schedule_pdf(
        self,
        schedule_id: int,
        group_ids: Optional[List[int]] = None,
        custom_filename: Optional[str] = None,
    ) -> Tuple[io.BytesIO, str]:
        """
        Export a schedule to PDF using the same data layout as Excel (landscape A4).

        Steps:
        - Load schedule and lessons (optionally filtered by groups).
        - Generate time slots and determine date range.
        - Build a ReportLab table with compact styling and colored cells.

        Args:
            schedule_id (int): ID of the schedule to export.
            group_ids (Optional[List[int]]): Subset of group IDs to include; None for all.
            custom_filename (Optional[str]): Base filename without extension; defaults to a generated name.

        Returns:
            Tuple[io.BytesIO, str]: (buffer, filename) where buffer is positioned at start.
        """
        """Экспорт расписания в PDF в том же формате что и Excel"""

        # Получаем расписание

        schedule = self.get_by_id(schedule_id)

        # Получаем уроки для расписания (с фильтрацией по группам если указаны)
        lessons = self._get_lessons_for_schedule(schedule_id, group_ids)

        # Создаем временные слоты (с 8:00 до 18:00 с интервалом 45 минут)
        time_slots = self._generate_time_slots()

        # Определяем период дат из уроков или семестра
        if lessons:
            start = min(lesson.date for lesson in lessons)
            end = max(lesson.date for lesson in lessons)
        else:
            # Если уроков нет, используем даты семестра
            start = schedule.semester.start_date
            end = schedule.semester.end_date

        # Получаем выходные дни для этого периода
        holiday_service = UniversityHolidayService(self.db)
        holiday_dates_list = holiday_service.get_expanded_holiday_dates(start, end)
        holiday_dates = {holiday["date"] for holiday in holiday_dates_list}

        # Создаем PDF
        output = io.BytesIO()

        # Используем landscape ориентацию для лучшего размещения таблицы
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            leftMargin=10 * mm,
            rightMargin=10 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
        )

        # Создаем содержимое PDF
        story = []

        # Заголовок на польском как в Excel
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=getSampleStyleSheet()["Heading1"],
            fontSize=9,  # Arial размер 9
            alignment=1,  # CENTER
            spaceAfter=8,  # Уменьшенный отступ
            fontName="Helvetica-Bold",  # Используем поддерживаемый шрифт
        )

        title_text = f"PLAN ZAJĘĆ - ROK AKADEMICKI {schedule.academic_year.name}"
        story.append(Paragraph(title_text, title_style))

        # Подзаголовок на польском как в Excel
        subtitle_style = ParagraphStyle(
            "CustomSubtitle",
            parent=getSampleStyleSheet()["Heading2"],
            fontSize=9,  # Arial размер 9
            alignment=1,  # CENTER
            spaceAfter=12,  # Уменьшенный отступ
            fontName="Helvetica-Bold",  # Используем поддерживаемый шрифт
        )

        subtitle_text = f"{schedule.direction.name} - {schedule.direction.faculty.name} - ROK {schedule.semester.number} SEMESTR"
        story.append(Paragraph(subtitle_text, subtitle_style))

        # Создаем таблицу данных
        table_data, cell_colors = self._create_pdf_table_data(
            start, end, time_slots, lessons, holiday_dates
        )

        # Создаем таблицу
        table = Table(table_data)

        # Стили таблицы (оптимизировано для landscape)
        table_style = TableStyle(
            [
                # Границы
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                # Заголовки временных слотов
                ("BACKGROUND", (1, 0), (-2, 0), colors.lightgrey),
                ("ALIGN", (1, 0), (-2, 0), "CENTER"),
                ("VALIGN", (1, 0), (-2, 0), "MIDDLE"),
                ("FONTNAME", (1, 0), (-2, 0), "Helvetica-Bold"),
                ("FONTSIZE", (1, 0), (-2, 0), 7),  # Уменьшенный размер для времени
                # Колонка GRUPA
                ("BACKGROUND", (-1, 0), (-1, 0), colors.lightgrey),
                ("ALIGN", (-1, 0), (-1, 0), "CENTER"),
                ("VALIGN", (-1, 0), (-1, 0), "MIDDLE"),
                ("FONTNAME", (-1, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (-1, 0), (-1, 0), 8),
                # Колонка дат
                ("BACKGROUND", (0, 1), (0, -1), colors.whitesmoke),
                ("ALIGN", (0, 1), (0, -1), "CENTER"),
                ("VALIGN", (0, 1), (0, -1), "MIDDLE"),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 1), (0, -1), 7),  # Уменьшенный размер для дат
                # Ячейки уроков
                ("ALIGN", (1, 1), (-2, -1), "CENTER"),
                ("VALIGN", (1, 1), (-2, -1), "MIDDLE"),
                ("FONTNAME", (1, 1), (-2, -1), "Helvetica"),
                ("FONTSIZE", (1, 1), (-2, -1), 9),  # Helvetica размер 9 обычный
                # Колонка групп
                ("ALIGN", (-1, 1), (-1, -1), "CENTER"),
                ("VALIGN", (-1, 1), (-1, -1), "MIDDLE"),
                ("FONTNAME", (-1, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (-1, 1), (-1, -1), 9),  # Helvetica размер 9 обычный
            ]
        )

        table.setStyle(table_style)

        # Применяем цвета к ячейкам с уроками
        for (row, col), hex_color in cell_colors.items():
            rgb_color = self._hex_to_rgb(hex_color)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (col, row), (col, row), rgb_color),
                        (
                            "TEXTCOLOR",
                            (col, row),
                            (col, row),
                            colors.black,
                        ),  # Черный текст как в Excel
                    ]
                )
            )

        # Создаем объединенные ячейки для многочасовых уроков и дат
        self._apply_pdf_cell_spans(table, table_data, lessons, time_slots)

        # Устанавливаем ширину колонок (оптимизировано для landscape A4)
        available_width = landscape(A4)[0] - 20 * mm  # Общая ширина минус отступы
        date_col_width = 25 * mm  # Колонка дат
        group_col_width = 20 * mm  # Колонка групп

        # Распределяем оставшуюся ширину между временными слотами
        remaining_width = available_width - date_col_width - group_col_width
        time_slot_width = remaining_width / len(time_slots)

        # Если слишком много временных слотов, уменьшаем ширину колонок
        if len(time_slots) > 10:
            # Уменьшаем колонки дат и групп для экономии места
            date_col_width = 20 * mm
            group_col_width = 15 * mm
            remaining_width = available_width - date_col_width - group_col_width
            time_slot_width = remaining_width / len(time_slots)

        col_widths = [date_col_width]  # Колонка дат
        col_widths.extend([time_slot_width] * len(time_slots))  # Временные слоты
        col_widths.append(group_col_width)  # Колонка групп
        table._argW = col_widths

        story.append(table)

        # Генерируем PDF
        doc.build(story)

        # Генерируем имя файла
        if custom_filename:
            filename = f"{custom_filename}.pdf"
        else:
            filename = f"schedule_{schedule.direction.name}_{schedule.semester.number}_{start.strftime('%Y-%m-%d')}_to_{end.strftime('%Y-%m-%d')}.pdf"
        filename = filename.replace(" ", "_").replace("/", "_")

        output.seek(0)
        return output, filename

    def _create_pdf_table_data(
        self, start_date, end_date, time_slots, lessons, holiday_dates=None
    ):
        """
        Build PDF table data and a map of cell colors.

        The returned table_data is a list of rows for ReportLab Table:
        - Row 0 is the header: ["Date/Time", ...time_slots..., "GRUPA"]
        - Subsequent rows represent days (one or multiple rows depending on groups)

        Args:
            start_date (date): Start date of the range.
            end_date (date): End date of the range.
            time_slots (list[str]): Time slot labels (e.g., "08:00-08:45").
            lessons (list[Lesson]): Lessons to render.

        Returns:
            tuple[list[list[str]], dict[tuple[int, int], str]]:
                (table_data, cell_colors) where cell_colors keys are (row_index, col_index)
                and values are hex colors for background.
        """
        """Создание данных для PDF таблицы"""
        table_data = []
        cell_colors = {}  # Словарь для хранения цветов ячеек: {(row, col): color}

        # Заголовочная строка (используем пустую ячейку как в Excel)
        header_row = [""] + time_slots + ["GRUPA"]
        table_data.append(header_row)

        current_date = start_date
        table_row = 1  # Начинаем с 1, так как 0-я строка - заголовок

        # Получаем расписание для контекста
        schedule = self.get_by_id(lessons[0].schedule_id if lessons else 1)

        while current_date <= end_date:
            # Получаем уроки на этот день
            day_lessons = [lesson for lesson in lessons if lesson.date == current_date]
            is_holiday = self._is_holiday(current_date, schedule, holiday_dates)

            if not day_lessons:
                # Пустая строка для дня без уроков
                date_text = f"{self._get_weekday_polish(current_date)}\n{current_date.strftime('%d.%m.%Y')}"
                row = [date_text] + [""] * len(time_slots) + [""]
                table_data.append(row)

                # Если это выходной день, отмечаем все ячейки серым цветом
                if is_holiday:
                    for col_idx in range(len(time_slots) + 2):  # +2 для даты и группы
                        cell_colors[(table_row, col_idx)] = "a6a6a6"

                table_row += 1
            else:
                # Определяем уникальные группы для этого дня (теперь у урока может быть несколько групп)
                groups_for_day = {}
                for lesson in day_lessons:
                    if lesson.groups:
                        for group in lesson.groups:
                            group_name = group.name
                            if group_name not in groups_for_day:
                                groups_for_day[group_name] = []
                            groups_for_day[group_name].append(lesson)
                    else:
                        # Урок без групп
                        if "No Group" not in groups_for_day:
                            groups_for_day["No Group"] = []
                        groups_for_day["No Group"].append(lesson)

                if len(groups_for_day) <= 1:
                    # Одна группа или уроки без группы
                    date_text = f"{self._get_weekday_polish(current_date)}\n{current_date.strftime('%d.%m.%Y')}"
                    row = [date_text]

                    # Заполняем временные слоты с учетом объединения для многочасовых уроков
                    processed_lessons = set()  # Отслеживаем обработанные уроки
                    for col_idx, time_slot in enumerate(time_slots):
                        lesson_in_slot = self._find_lesson_in_slot(
                            day_lessons, time_slot
                        )
                        if (
                            lesson_in_slot
                            and lesson_in_slot.id not in processed_lessons
                        ):
                            cell_text = self._format_lesson_for_pdf(lesson_in_slot)
                            # Сохраняем цвет для этой ячейки
                            if lesson_in_slot.professor.professor_profile.color:
                                # Вычисляем сколько слотов занимает урок
                                slots_count = self._get_lesson_time_slots_count(
                                    lesson_in_slot, time_slots
                                )
                                # Сохраняем цвет для всех ячеек урока
                                for slot_offset in range(slots_count):
                                    if col_idx + slot_offset < len(time_slots):
                                        cell_colors[
                                            (table_row, col_idx + 1 + slot_offset)
                                        ] = lesson_in_slot.professor.professor_profile.color
                            processed_lessons.add(lesson_in_slot.id)
                        elif lesson_in_slot and lesson_in_slot.id in processed_lessons:
                            # Урок уже обработан, пустая ячейка для объединения
                            cell_text = ""
                        else:
                            cell_text = ""
                        row.append(cell_text)

                    # Колонка групп
                    groups_text = (
                        "\n".join(sorted(groups_for_day.keys()))
                        if groups_for_day
                        else ""
                    )
                    row.append(groups_text)
                    table_data.append(row)

                    # Отмечаем выходной день серым цветом
                    if is_holiday:
                        for col_idx in range(
                            len(time_slots) + 2
                        ):  # +2 для даты и группы
                            cell_colors[(table_row, col_idx)] = "a6a6a6"

                    table_row += 1
                else:
                    # Несколько групп - отдельные строки для каждой группы
                    num_groups = len(groups_for_day)
                    day_end_row = table_row + num_groups - 1

                    first_group = True
                    for group_name in sorted(groups_for_day.keys()):
                        group_lessons = groups_for_day[group_name]

                        if first_group:
                            date_text = f"{self._get_weekday_polish(current_date)}\n{current_date.strftime('%d.%m.%Y')}"
                            first_group = False
                        else:
                            date_text = ""

                        row = [date_text]

                        # Заполняем временные слоты для этой группы с учетом объединения
                        processed_lessons = set()
                        for col_idx, time_slot in enumerate(time_slots):
                            lesson_in_slot = self._find_lesson_in_slot(
                                group_lessons, time_slot
                            )
                            if (
                                lesson_in_slot
                                and lesson_in_slot.id not in processed_lessons
                            ):
                                cell_text = self._format_lesson_for_pdf(lesson_in_slot)
                                # Сохраняем цвет для этой ячейки
                                if lesson_in_slot.professor.professor_profile.color:
                                    slots_count = self._get_lesson_time_slots_count(
                                        lesson_in_slot, time_slots
                                    )
                                    for slot_offset in range(slots_count):
                                        if col_idx + slot_offset < len(time_slots):
                                            cell_colors[
                                                (table_row, col_idx + 1 + slot_offset)
                                            ] = lesson_in_slot.professor.professor_profile.color
                                processed_lessons.add(lesson_in_slot.id)
                            elif (
                                lesson_in_slot
                                and lesson_in_slot.id in processed_lessons
                            ):
                                cell_text = ""
                            else:
                                cell_text = ""
                            row.append(cell_text)

                        # Колонка групп
                        row.append(group_name)
                        table_data.append(row)
                        table_row += 1

                    # Отмечаем выходной день серым цветом для всех строк групп
                    if is_holiday:
                        for row_idx in range(day_row_start, table_row):
                            for col_idx in range(
                                len(time_slots) + 2
                            ):  # +2 для даты и группы
                                cell_colors[(row_idx, col_idx)] = "a6a6a6"

            current_date += timedelta(days=1)

        return table_data, cell_colors

    def _hex_to_rgb(self, hex_color: str):
        """
        Convert a hex color string into a ReportLab Color (0..1 per channel).

        Args:
            hex_color (str): Hex color in the form "#RRGGBB" or "RRGGBB".

        Returns:
            reportlab.lib.colors.Color: RGB color for ReportLab.
        """
        """Конвертация hex цвета в RGB для ReportLab"""
        # Убираем # если есть
        hex_color = hex_color.lstrip("#")

        # Конвертируем в RGB (значения от 0 до 1)
        r = int(hex_color[0:2], 16) / 255.0
        g = int(hex_color[2:4], 16) / 255.0
        b = int(hex_color[4:6], 16) / 255.0

        return colors.Color(r, g, b)

    def _is_holiday(self, date_obj, schedule=None, holiday_dates=None) -> bool:
        """
        Check if a date is a university holiday.

        Args:
            date_obj (date): Date to check.
            schedule (Schedule): Schedule object for context (optional if holiday_dates provided).
            holiday_dates (set): Pre-calculated holiday dates (optional).

        Returns:
            bool: True if the date is a holiday.
        """
        if holiday_dates is not None:
            return date_obj in holiday_dates

        # Используем только get_expanded_holiday_dates для получения выходных дней
        try:
            if not schedule:
                return False

            # Используем UniversityHolidayService для получения выходных дней
            holiday_service = UniversityHolidayService(self.db)
            semester_start = schedule.semester.start_date
            semester_end = schedule.semester.end_date

            holiday_dates_list = holiday_service.get_expanded_holiday_dates(
                semester_start, semester_end
            )

            # Проверяем, есть ли наша дата среди выходных
            return any(holiday["date"] == date_obj for holiday in holiday_dates_list)

        except Exception:
            return False

    def _apply_pdf_cell_spans(self, table, table_data, lessons, time_slots):
        """
        Apply cell spans for multi-hour lessons and date merging in PDF table.

        Args:
            table (Table): ReportLab table object.
            table_data (list): Table data structure.
            lessons (list): List of lessons.
            time_slots (list): List of time slots.
        """
        """Применяем объединение ячеек для многочасовых уроков и дат в PDF"""

        # Сначала обрабатываем объединение дат для дней с несколькими группами
        current_date_text = None
        date_start_row = None

        for row_idx in range(1, len(table_data)):
            date_cell = table_data[row_idx][0]  # Первая колонка - даты

            if date_cell and date_cell != current_date_text:
                # Завершаем предыдущее объединение если было
                if (
                    current_date_text
                    and date_start_row is not None
                    and row_idx - date_start_row > 1
                ):
                    table.setStyle(
                        TableStyle(
                            [
                                ("SPAN", (0, date_start_row), (0, row_idx - 1)),
                            ]
                        )
                    )

                # Начинаем новую дату
                current_date_text = date_cell
                date_start_row = row_idx
            elif not date_cell and current_date_text:
                # Пустая ячейка даты - продолжение предыдущей даты
                continue

        # Завершаем последнее объединение если нужно
        if (
            current_date_text
            and date_start_row is not None
            and len(table_data) - date_start_row > 1
        ):
            table.setStyle(
                TableStyle(
                    [
                        ("SPAN", (0, date_start_row), (0, len(table_data) - 1)),
                    ]
                )
            )

        # Теперь обрабатываем объединение многочасовых уроков
        for row_idx in range(1, len(table_data)):
            processed_lessons = set()

            # Проходим по всем временным слотам
            for col_idx in range(
                1, len(time_slots) + 1
            ):  # +1 потому что первая колонка - дата
                cell_content = table_data[row_idx][col_idx]

                # Если ячейка не пустая и содержит урок
                if cell_content and cell_content not in processed_lessons:
                    # Ищем урок по содержимому ячейки
                    lesson = self._find_lesson_by_content(lessons, cell_content)
                    if lesson:
                        slots_count = self._get_lesson_time_slots_count(
                            lesson, time_slots
                        )

                        if slots_count > 1:
                            # Создаем объединение ячеек
                            start_col = col_idx
                            end_col = min(col_idx + slots_count - 1, len(time_slots))

                            # Применяем стиль объединения через SPAN
                            table.setStyle(
                                TableStyle(
                                    [
                                        (
                                            "SPAN",
                                            (start_col, row_idx),
                                            (end_col, row_idx),
                                        ),
                                    ]
                                )
                            )

                        processed_lessons.add(cell_content)

    def _find_lesson_by_content(self, lessons, content):
        """
        Find lesson by formatted content string.

        Args:
            lessons (list): List of lessons.
            content (str): Formatted lesson content.

        Returns:
            Lesson or None: Matching lesson if found.
        """
        """Поиск урока по отформатированному содержимому"""
        for lesson in lessons:
            formatted_content = self._format_lesson_for_pdf(lesson)
            if formatted_content == content:
                return lesson
        return None

    def _format_lesson_for_pdf(self, lesson) -> str:
        """
        Produce a compact, multi-line text representation of a lesson for PDF cells.

        Format:
        - Subject name with lesson type
        - Professor full name with title
        - Delivery format ("online" or room number)

        Args:
            lesson (Lesson): Lesson to render.

        Returns:
            str: Compact string for the PDF cell.
        """
        """Форматирование урока для PDF (компактная версия)"""
        # Формируем название предмета с типом
        lesson_text = lesson.subject.name
        if hasattr(lesson, "lesson_type") and lesson.lesson_type:
            lesson_text += f" - {lesson.lesson_type}"

        # Сокращаем для PDF если слишком длинное
        if len(lesson_text) > 25:
            lesson_text = lesson_text[:22] + "..."

        # Добавляем полное имя преподавателя со званием
        professor_full = f"{lesson.professor.name} {lesson.professor.surname}"
        if hasattr(lesson.professor, "title") and lesson.professor.title:
            professor_full = f"{lesson.professor.title} {professor_full}"

        # Сокращаем имя преподавателя для PDF если слишком длинное
        if len(professor_full) > 25:
            professor_full = professor_full[:22] + "..."

        lesson_text += f"\n{professor_full}"

        # Добавляем информацию о формате проведения
        if lesson.is_online:
            lesson_text += "\nonline"
        elif lesson.room:
            lesson_text += f"\n{lesson.room.number}"

        return lesson_text

    def _get_lessons_for_schedule(
        self, schedule_id: int, group_ids: Optional[List[int]] = None
    ) -> List[Lesson]:
        """
        Fetch lessons for a schedule, optionally restricted to specific groups.

        Args:
            schedule_id (int): Schedule ID.
            group_ids (Optional[List[int]]): List of group IDs to include; None for all.

        Returns:
            List[Lesson]: Lessons ordered by date and start time.
        """
        """Получение уроков для расписания с возможной фильтрацией по группам"""
        from ..models.group import Group

        query = self.db.query(Lesson).filter(Lesson.schedule_id == schedule_id)

        # Если указаны конкретные группы, фильтруем по ним через many-to-many
        if group_ids:
            query = query.join(Lesson.groups).filter(Group.id.in_(group_ids))

        return query.order_by(Lesson.date, Lesson.start_time).all()

    def _get_groups_for_schedule(
        self, schedule_id: int, group_ids: Optional[List[int]] = None
    ) -> List[Group]:
        """
        Fetch groups for a schedule, optionally restricted to specific groups.

        Uses the schedule's direction and semester to scope groups.

        Args:
            schedule_id (int): Schedule ID.
            group_ids (Optional[List[int]]): List of group IDs to include; None for all.

        Returns:
            List[Group]: Groups ordered by name.
        """
        """Получение групп для расписания с возможной фильтрацией"""
        schedule = self.db.query(Schedule).filter(Schedule.id == schedule_id).first()

        query = (
            self.db.query(Group)
            .join(StudyForm)
            .filter(
                and_(
                    StudyForm.direction_id == schedule.direction_id,
                    Group.semester_id == schedule.semester_id,
                )
            )
        )

        # Если указаны конкретные группы, фильтруем по ним
        if group_ids:
            query = query.filter(Group.id.in_(group_ids))

        return query.order_by(Group.name).all()

    def _generate_time_slots(self) -> List[str]:
        """
        Generate time slot labels from 08:00 to 21:15.

        Each slot is 45 minutes followed by a 5-minute break for academic hours.

        Returns:
            List[str]: Sequence of time slot labels.
        """
        """Генерация временных слотов"""
        slots = []
        start_time = datetime.strptime("08:00", "%H:%M")
        end_time = datetime.strptime("21:15", "%H:%M")

        current = start_time
        while current < end_time:
            next_time = current + timedelta(minutes=45)
            slot = f"{current.strftime('%H:%M')}-{next_time.strftime('%H:%M')}"
            slots.append(slot)
            current = next_time + timedelta(
                minutes=5
            )  # 5 минут перерыв между академическими часами

        return slots

    def _setup_styles(
        self,
        worksheet,
        schedule,
        start_date,
        end_date,
        time_slots,
        groups,
        lessons,
        holiday_dates=None,
    ):
        """
        Apply styles and fill the Excel worksheet with schedule content.

        Writes:
        - Title and subtitle (merged cells)
        - Time slot headers
        - Rows for each date (and per-group rows if multiple groups on the same day)
        - Lesson cells with subject color backgrounds and wrapped text

        Args:
            worksheet (openpyxl.worksheet.worksheet.Worksheet): Active sheet.
            schedule (Schedule): Schedule metadata (title/subtitle context).
            start_date (date): Start date for rendering.
            end_date (date): End date for rendering.
            time_slots (list[str]): Column headers for time intervals.
            groups (list[Group]): Not used in current implementation (reserved).
            lessons (list[Lesson]): Lessons to render in the grid.

        Returns:
            None
        """
        """Настройка стилей и заполнение данных"""

        # Стили Arial размер 9 для всей таблицы
        header_font = Font(name="Arial", bold=True, size=9, color="000000")
        title_font = Font(name="Arial", bold=True, size=9, color="000000")
        cell_font = Font(
            name="Arial", size=9, color="000000"
        )  # Обычный текст для групп и занятий
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Заголовок на польском (строка 1)
        title_text = f"PLAN ZAJĘĆ - ROK AKADEMICKI {schedule.academic_year.name}"
        worksheet.merge_cells(f"A1:{get_column_letter(len(time_slots) + 2)}1")
        title_cell = worksheet.cell(row=1, column=1, value=title_text)
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Подзаголовок на польском (строка 2)
        subtitle_text = f"{schedule.direction.name} - {schedule.direction.faculty.name} - ROK {schedule.semester.number} SEMESTR"
        worksheet.merge_cells(f"A2:{get_column_letter(len(time_slots) + 2)}2")
        subtitle_cell = worksheet.cell(row=2, column=1, value=subtitle_text)
        subtitle_cell.font = header_font
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Временные слоты (строка 3)
        for col, time_slot in enumerate(time_slots, 2):
            cell = worksheet.cell(row=3, column=col, value=time_slot)
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Колонка "GRUPA" (строка 3, последняя колонка)
        group_col = len(time_slots) + 2
        group_cell = worksheet.cell(row=3, column=group_col, value="GRUPA")
        group_cell.font = header_font
        group_cell.border = border
        group_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Закрепляем первые 3 строки (заголовки и временные слоты)
        worksheet.freeze_panes = "A4"

        # Заполнение дней и уроков
        current_date = start_date
        row = 4

        while current_date <= end_date:
            # Получаем уроки на этот день
            day_lessons = [lesson for lesson in lessons if lesson.date == current_date]

            # Проверяем, является ли день выходным
            is_holiday = holiday_dates and current_date in holiday_dates

            if not day_lessons:
                # Если нет уроков, создаем пустую строку
                date_text = f"{self._get_weekday_polish(current_date)}\n{current_date.strftime('%d.%m.%Y')}"
                date_cell = worksheet.cell(row=row, column=1, value=date_text)
                date_cell.font = header_font
                date_cell.border = border
                date_cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

                # Проверяем, является ли день выходным
                is_holiday = self._is_holiday(current_date, schedule, holiday_dates)

                # Заполняем пустые временные слоты
                for col, time_slot in enumerate(time_slots, 2):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = Font(
                        name="Arial", size=9, color="000000"
                    )  # Применяем базовый шрифт
                    cell.border = border

                    # Если это выходной день, закрашиваем ячейку серым
                    if is_holiday:
                        cell.fill = PatternFill(
                            start_color="a6a6a6", end_color="a6a6a6", fill_type="solid"
                        )

                # Пустая колонка групп
                groups_cell = worksheet.cell(row=row, column=group_col, value="")
                groups_cell.border = border

                # Если это выходной день, закрашиваем ячейку даты и группы серым
                if is_holiday:
                    date_cell.fill = PatternFill(
                        start_color="a6a6a6", end_color="a6a6a6", fill_type="solid"
                    )
                    groups_cell.fill = PatternFill(
                        start_color="a6a6a6", end_color="a6a6a6", fill_type="solid"
                    )

                row += 1
            else:
                # Определяем уникальные группы для этого дня (теперь у урока может быть несколько групп)
                groups_for_day = {}
                for lesson in day_lessons:
                    if lesson.groups:
                        for group in lesson.groups:
                            group_name = group.name
                            if group_name not in groups_for_day:
                                groups_for_day[group_name] = []
                            groups_for_day[group_name].append(lesson)
                    else:
                        # Урок без групп
                        if "No Group" not in groups_for_day:
                            groups_for_day["No Group"] = []
                        groups_for_day["No Group"].append(lesson)

                # Если есть только одна группа или уроки без группы
                if len(groups_for_day) <= 1:
                    # Стандартная строка
                    date_text = f"{self._get_weekday_polish(current_date)}\n{current_date.strftime('%d.%m.%Y')}"
                    date_cell = worksheet.cell(row=row, column=1, value=date_text)
                    date_cell.font = header_font
                    date_cell.border = border
                    date_cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

                    # Заполняем временные слоты с объединением для многочасовых уроков
                    processed_lessons = set()  # Отслеживаем обработанные уроки
                    for col, time_slot in enumerate(time_slots, 2):
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = Font(
                            name="Arial", size=9, color="000000"
                        )  # Применяем базовый шрифт
                        cell.border = border

                        # Ищем урок в этом временном слоте
                        lesson_in_slot = self._find_lesson_in_slot(
                            day_lessons, time_slot
                        )
                        if (
                            lesson_in_slot
                            and lesson_in_slot.id not in processed_lessons
                        ):
                            # Вычисляем сколько слотов занимает урок
                            slots_count = self._get_lesson_time_slots_count(
                                lesson_in_slot, time_slots
                            )

                            if slots_count > 1:
                                # Объединяем ячейки для многочасового урока
                                start_col = col
                                end_col = col + slots_count - 1
                                try:
                                    worksheet.merge_cells(
                                        f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}"
                                    )
                                    # Заполняем только первую ячейку
                                    self._fill_lesson_cell(cell, lesson_in_slot)
                                except:
                                    # Если объединение не удалось, просто заполняем текущую ячейку
                                    self._fill_lesson_cell(cell, lesson_in_slot)
                            else:
                                # Обычный урок на один слот
                                self._fill_lesson_cell(cell, lesson_in_slot)

                            processed_lessons.add(lesson_in_slot.id)

                    # Заполняем колонку групп
                    groups_text = (
                        "\n".join(sorted(groups_for_day.keys()))
                        if groups_for_day
                        else ""
                    )
                    groups_cell = worksheet.cell(
                        row=row, column=group_col, value=groups_text
                    )
                    groups_cell.font = cell_font
                    groups_cell.border = border
                    groups_cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

                    row += 1
                else:
                    # Несколько групп - создаем отдельные строки для каждой группы
                    start_row = row
                    num_groups = len(groups_for_day)

                    # Создаем объединенную ячейку для даты
                    date_text = f"{self._get_weekday_polish(current_date)}\n{current_date.strftime('%d.%m.%Y')}"
                    worksheet.merge_cells(f"A{start_row}:A{start_row + num_groups - 1}")
                    date_cell = worksheet.cell(row=start_row, column=1, value=date_text)
                    date_cell.font = header_font
                    date_cell.border = border
                    date_cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

                    # Создаем строки для каждой группы
                    for group_name in sorted(groups_for_day.keys()):
                        group_lessons = groups_for_day[group_name]

                        # Заполняем временные слоты только для этой группы с объединением
                        processed_lessons = set()
                        for col, time_slot in enumerate(time_slots, 2):
                            cell = worksheet.cell(row=row, column=col)
                            cell.font = Font(
                                name="Arial", size=9, color="000000"
                            )  # Применяем базовый шрифт
                            cell.border = border

                            # Ищем урок в этом временном слоте для этой группы
                            lesson_in_slot = self._find_lesson_in_slot(
                                group_lessons, time_slot
                            )
                            if (
                                lesson_in_slot
                                and lesson_in_slot.id not in processed_lessons
                            ):
                                # Вычисляем сколько слотов занимает урок
                                slots_count = self._get_lesson_time_slots_count(
                                    lesson_in_slot, time_slots
                                )

                                if slots_count > 1:
                                    # Объединяем ячейки для многочасового урока
                                    start_col = col
                                    end_col = col + slots_count - 1
                                    try:
                                        worksheet.merge_cells(
                                            f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}"
                                        )
                                        # Заполняем только первую ячейку
                                        self._fill_lesson_cell(cell, lesson_in_slot)
                                    except:
                                        # Если объединение не удалось, просто заполняем текущую ячейку
                                        self._fill_lesson_cell(cell, lesson_in_slot)
                                else:
                                    # Обычный урок на один слот
                                    self._fill_lesson_cell(cell, lesson_in_slot)

                                processed_lessons.add(lesson_in_slot.id)

                        # Заполняем колонку групп
                        groups_cell = worksheet.cell(
                            row=row, column=group_col, value=group_name
                        )
                        groups_cell.font = cell_font
                        groups_cell.border = border
                        groups_cell.alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )

                        row += 1

            current_date += timedelta(days=1)

        # Устанавливаем высоту 48 для всех строк данных (начиная с 4-й строки)
        for row_num in range(4, row):
            worksheet.row_dimensions[row_num].height = 48

    def _get_weekday_polish(self, date_obj) -> str:
        """
        Get Polish weekday name for the given date.

        Args:
            date_obj (date): Date object.

        Returns:
            str: Polish weekday name.
        """
        polish_weekdays = {
            0: "PONIEDZIALEK",  # Monday (убираем диакритические знаки для PDF)
            1: "WTOREK",  # Tuesday
            2: "SRODA",  # Wednesday
            3: "CZWARTEK",  # Thursday
            4: "PIATEK",  # Friday
            5: "SOBOTA",  # Saturday
            6: "NIEDZIELA",  # Sunday
        }
        return polish_weekdays.get(date_obj.weekday(), "NIEZNANY")

    def _get_weekday_english(self, date_obj) -> str:
        """
        Get the English uppercase weekday name for a date.

        Args:
            date_obj (date): Target date.

        Returns:
            str: Weekday name (e.g., "MONDAY").
        """
        """Получение дня недели на английском"""
        weekdays = {
            0: "MONDAY",
            1: "TUESDAY",
            2: "WEDNESDAY",
            3: "THURSDAY",
            4: "FRIDAY",
            5: "SATURDAY",
            6: "SUNDAY",
        }
        return weekdays[date_obj.weekday()]

    def _find_lesson_in_slot(
        self, lessons: List[Lesson], time_slot: str
    ) -> Optional[Lesson]:
        """
        Find a lesson that overlaps a given time slot.

        Overlap conditions (inclusive):
        - Lesson starts before slot end and ends after slot start, or
        - Either endpoint falls within the other interval, or
        - One interval fully contains the other.

        Args:
            lessons (List[Lesson]): Candidate lessons for a specific date.
            time_slot (str): Slot label "HH:MM-HH:MM".

        Returns:
            Optional[Lesson]: Matching lesson if found, otherwise None.
        """
        """Поиск урока в временном слоте"""
        slot_start, slot_end = time_slot.split("-")
        slot_start_time = datetime.strptime(slot_start, "%H:%M").time()
        slot_end_time = datetime.strptime(slot_end, "%H:%M").time()

        for lesson in lessons:
            # Проверяем пересечение времени
            if (
                lesson.start_time <= slot_start_time < lesson.end_time
                or lesson.start_time < slot_end_time <= lesson.end_time
                or (
                    lesson.start_time >= slot_start_time
                    and lesson.end_time <= slot_end_time
                )
            ):
                return lesson

        return None

    def _get_lesson_time_slots_count(
        self, lesson: Lesson, time_slots: List[str]
    ) -> int:
        """
        Calculate how many time slots a lesson spans.

        Args:
            lesson (Lesson): The lesson to check.
            time_slots (List[str]): Available time slots.

        Returns:
            int: Number of time slots the lesson spans.
        """
        lesson_start = lesson.start_time
        lesson_end = lesson.end_time

        # Convert times to minutes for easier calculation
        lesson_start_minutes = lesson_start.hour * 60 + lesson_start.minute
        lesson_end_minutes = lesson_end.hour * 60 + lesson_end.minute

        # Each time slot is 45 minutes + 5 minutes break = 50 minutes
        # But we need to check how many slots the lesson actually occupies
        slots_count = 0
        for slot in time_slots:
            slot_start_str, slot_end_str = slot.split("-")
            slot_start = datetime.strptime(slot_start_str, "%H:%M").time()
            slot_end = datetime.strptime(slot_end_str, "%H:%M").time()

            slot_start_minutes = slot_start.hour * 60 + slot_start.minute
            slot_end_minutes = slot_end.hour * 60 + slot_end.minute

            # Check if lesson overlaps with this slot
            if (
                lesson_start_minutes < slot_end_minutes
                and lesson_end_minutes > slot_start_minutes
            ):
                slots_count += 1

        return max(1, slots_count)

    def _fill_lesson_cell(self, cell, lesson: Lesson):
        """
        Fill an Excel cell with lesson data and apply styling.

        Content:
        - Subject name
        - Lesson type (if available)
        - Professor full name with title (if available)
        - Delivery: "online" or room number

        Styling:
        - Center alignment, black font
        - Background color from professor color (hex) or gray fallback

        Args:
            cell (openpyxl.cell.Cell): Target cell to fill.
            lesson (Lesson): Lesson providing values and color.

        Returns:
            None
        """
        """Заполнение ячейки урока"""
        # Формируем текст урока
        lesson_text = lesson.subject.name

        # Добавляем тип урока если есть
        if hasattr(lesson, "lesson_type") and lesson.lesson_type:
            lesson_text += f" - {lesson.lesson_type}"

        # Добавляем полное имя преподавателя со званием
        professor_full = f"{lesson.professor.name} {lesson.professor.surname}"
        if hasattr(lesson.professor, "title") and lesson.professor.title:
            professor_full = f"{lesson.professor.title} {professor_full}"
        lesson_text += f" - {professor_full}"

        # Добавляем информацию о формате проведения
        if lesson.is_online:
            lesson_text += "\nonline"
        elif lesson.room:
            lesson_text += f"\n{lesson.room.number}"

        cell.value = lesson_text
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.font = Font(
            name="Arial", size=9, color="000000", bold=False
        )  # Черный обычный текст Arial

        # Устанавливаем цвет фона из цвета преподавателя
        if lesson.professor.professor_profile.color:
            # Убираем # из hex цвета
            color_hex = lesson.professor.professor_profile.color.replace("#", "")
            cell.fill = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type="solid"
            )
        else:
            # Цвет по умолчанию (черный)
            cell.fill = PatternFill(
                start_color="000000", end_color="000000", fill_type="solid"
            )
